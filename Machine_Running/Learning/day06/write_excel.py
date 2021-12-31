from openpyxl import Workbook

# 데이터 변수 선언
filename = 'sample.xlsx'

# 엑셀 파일 생성
wb = Workbook()  # 엑셀 파일 생성, Sheet1 자동 생성
ws = wb.active  # 시트 활성화
ws.title = 'new sheet'  # 시트 명 변경
ws['A1'] = 'Language'  # 시트 데이터 삽입
ws['B1'] = 'Create'

ws2 = wb.create_sheet('data_sheet')  # 2번째 시트 생성
ws2.cell(1, 1).value = 'Name'
ws2.cell(1, 2).value = 'Phone'
ws2.cell(2, 1).value = 'Tom'
ws2.cell(2, 2).value = '010-1111-2222'
wb.active = ws2
wb.save(filename=filename)  # 엑셀 파일 생성

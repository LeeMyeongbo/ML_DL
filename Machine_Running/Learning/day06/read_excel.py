import openpyxl

# 데이터 변수 선언
filename = 'sample.xlsx'

# 엑셀 파일 열기
book = openpyxl.load_workbook(filename)

# 시트 추출
sheet = book.worksheets[1]  # 1번 index의 sheet 가져옴
sheetnames = book.sheetnames  # 여러 시트들의 이름을 리스트로 저장
print('sheet => ' + sheet.title)
print('sheetnames =>', sheetnames)
# 1번 index의 시트의 행과 열 수를 출력
print('sheet.max_row = {}, sheet.nmax_column = {}'.format(sheet.max_row, sheet.max_column))

rowdata = []
for row in sheet.rows:
    rowdata.append([
        row[0].value,
        row[sheet.max_column-1].value
    ])
print(rowdata)
print('len(rowdata) => ' + str(len(rowdata)))
